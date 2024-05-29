import openpyxl


def save_to_excel(file_name, scratch_price, crushed_price, scratch_area, crushed_area, scratch_area_percent,
                  crushed_area_percent):
    xlsx_path = '/home/t24120/carkey_v1.0/backend/src/main/ai/price_predict.xlsx'

    workbook = openpyxl.load_workbook(xlsx_path)
    sheet = workbook.active

    def find_last_data_row(sheet):
        for row in range(sheet.max_row, 0, -1):
            if any(cell.value is not None for cell in sheet[row]):
                return row
        return 1

    # 마지막 행 찾기
    last_data_row = find_last_data_row(sheet)
    last_row = last_data_row + 1

    # 입력할 데이터 준비
    data = [file_name, scratch_price, crushed_price, scratch_area, crushed_area, scratch_area_percent,
            crushed_area_percent]

    # 데이터 엑셀 파일에 쓰기
    for index, value in enumerate(data, start=1):
        sheet.cell(row=last_row, column=index, value=value)

    # 변경사항 저장
    workbook.save(xlsx_path)
